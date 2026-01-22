from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
import app_config as cfg  # Asegúrate de que este nombre coincida con tu archivo de configuración

def aplicar_maquillaje(ws, df_ref, listas, col_map):
    print(f"Estilizando {ws.title}...")
    for merge in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merge))
    
    lista_grupos = listas['grupos']; lista_com = listas['comments']
    lista_reglas = listas['reglas_dos']; lista_dmd = listas.get('filas_dmd', [])
    ci_f = col_map['inicio_final']; cf_f = col_map['fin_final']
    ci = col_map['inicio']; cf = col_map['fin']

    ws.insert_cols(5); ws.cell(1, 5).value = ""; ws.cell(2, 5).value = "Comments"
    align = Alignment(horizontal='left', vertical='top')
    
    # CORRECCIÓN: Definición de colores sólida (start y end iguales)
    y_fill = PatternFill(start_color=cfg.COLOR_AMARILLO[0], end_color=cfg.COLOR_AMARILLO[0], fill_type="solid")
    font_y = Font(color=cfg.COLOR_AMARILLO[0]); font_w = Font(color="FFFFFF")
    com_map = {i['row']: i['status'] for i in lista_com}
    
    # 1. Alineación y Faux-Merge
    for ini, fin in lista_grupos:
        st = com_map.get(ini, ""); is_alert = st in ["OOS", "USTN"]
        for c in range(1, 6):
            ws.cell(ini, c).alignment = align
            if c == 5: 
                ws.cell(ini, c).value = st
                if is_alert: ws.cell(ini, c).fill = y_fill
            for r in range(ini + 1, fin + 1):
                cell = ws.cell(r, c); cell.alignment = align
                if c < 5: cell.font = font_w
                elif c == 5:
                    cell.value = st
                    if is_alert: cell.fill = y_fill; cell.font = font_y
                    else: cell.font = font_w

    # 2. Estilos específicos
    fill_n = PatternFill(start_color=cfg.COLOR_NAVY[0], end_color=cfg.COLOR_NAVY[0], fill_type="solid")
    font_n = Font(color=cfg.COLOR_NAVY[1])
    for r in lista_dmd:
        for c in range(ci_f, cf_f + 1): 
            if c != 5: ws.cell(r, c).fill = fill_n; ws.cell(r, c).font = font_n

    fill_k = PatternFill(start_color=cfg.COLOR_WK01_STATIC[0], end_color=cfg.COLOR_WK01_STATIC[0], fill_type="solid")
    font_k = Font(color=cfg.COLOR_WK01_STATIC[1])
    for rule in lista_reglas:
        r = rule['row_actual']; cell = ws.cell(r, ci_f)
        cell.fill = fill_k; cell.font = font_k; cell.number_format = cfg.FORMATO_CONTABILIDAD_DECIMAL
        for c in range(ci_f + 1, cf_f + 1): 
            if c != 5: ws.cell(r, c).number_format = cfg.FORMATO_CONTABILIDAD_DECIMAL

    for c in range(ci_f, cf_f + 1): 
        if c != 5: ws.cell(1, c).number_format = cfg.FORMATO_FECHA
    for r in range(1, 3):
        for c in range(1, ws.max_column + 1): ws.cell(r, c).font = Font(bold=True)

    # 3. Resumen
    last = df_ref.shape[0]; start_res = last + 5
    meds = ["Dmd Total", "Exec WSD", "Exec WCD", "INV"]
    
    for c_orig in range(ci, cf + 1):
        c_fin = c_orig + 1 + (1 if c_orig + 1 >= 5 else 0)
        if c_fin == 5: continue
        val = ws.cell(2, c_fin).value
        ws.cell(start_res - 1, c_fin, value=val).font = Font(bold=True)

    for i, m in enumerate(meds):
        r = start_res + i
        ws.cell(r, 6, value=m).font = Font(bold=True)
        for c_orig in range(ci, cf + 1):
            c_fin = c_orig + 1 + (1 if c_orig + 1 >= 5 else 0)
            if c_fin == 5: continue
            L = get_column_letter(c_fin)
            f = f"=SUMPRODUCT(SUBTOTAL(109,OFFSET({L}3,ROW({L}$3:{L}${last})-ROW({L}3),,1)),--($F$3:$F${last}=\"{m}\"))"
            ws.cell(r, c_fin, value=f).number_format = cfg.FORMATO_CONTABILIDAD_ENTERO

    r_doi = start_res + 4; r_tgt = start_res + 5
    ws.cell(r_doi, 6, value="Actual DOI").font = Font(bold=True)
    ws.cell(r_tgt, 6, value="DOI Target").font = Font(bold=True)

    for c_orig in range(ci, cf + 1):
        c_fin = c_orig + 1 + (1 if c_orig + 1 >= 5 else 0)
        if c_fin == 5: continue
        L = get_column_letter(c_fin); Ln = get_column_letter(c_fin + 1); Le = get_column_letter(c_fin + 8)
        f = f"=IFERROR(({L}{start_res+3}/(SUM({Ln}{start_res}:{Le}{start_res})/8))*7, 0)"
        ws.cell(r_doi, c_fin, value=f).number_format = cfg.FORMATO_CONTABILIDAD_DECIMAL
        ws.cell(r_tgt, c_fin, value=30).number_format = cfg.FORMATO_CONTABILIDAD_DECIMAL

    # --- CORRECCIÓN CRÍTICA: Definición de Estilos Condicionales ---
    st = get_column_letter(ci_f); en = get_column_letter(cf_f)
    rng = f"{st}{r_doi}:{en}{r_doi}"
    
    # Importante: end_color debe estar definido para compatibilidad
    s_p = DifferentialStyle(fill=PatternFill(start_color=cfg.COLOR_R_MORADO, end_color=cfg.COLOR_R_MORADO, fill_type="solid"))
    s_g = DifferentialStyle(fill=PatternFill(start_color=cfg.COLOR_R_VERDE, end_color=cfg.COLOR_R_VERDE, fill_type="solid"))
    s_y = DifferentialStyle(fill=PatternFill(start_color=cfg.COLOR_R_AMARILLO, end_color=cfg.COLOR_R_AMARILLO, fill_type="solid"))
    s_r = DifferentialStyle(fill=PatternFill(start_color=cfg.COLOR_R_ROJO, end_color=cfg.COLOR_R_ROJO, fill_type="solid"))
    
    ws.conditional_formatting.add(rng, Rule(type="cellIs", operator="equal", formula=[0], stopIfTrue=True))
    ws.conditional_formatting.add(rng, Rule(type="cellIs", operator="greaterThan", formula=[70], dxf=s_p))
    ws.conditional_formatting.add(rng, Rule(type="cellIs", operator="between", formula=[30, 70], dxf=s_g))
    ws.conditional_formatting.add(rng, Rule(type="cellIs", operator="between", formula=[15, 30], dxf=s_y))
    ws.conditional_formatting.add(rng, Rule(type="cellIs", operator="lessThan", formula=[15], dxf=s_r))

    s_dos_p = DifferentialStyle(font=Font(color=cfg.COLOR_MORADO[1]), fill=PatternFill(start_color=cfg.COLOR_MORADO[0], end_color=cfg.COLOR_MORADO[0], fill_type="solid"))
    s_dos_r = DifferentialStyle(font=Font(color=cfg.COLOR_ROJO[1]), fill=PatternFill(start_color=cfg.COLOR_ROJO[0], end_color=cfg.COLOR_ROJO[0], fill_type="solid"))
    s_dos_g = DifferentialStyle(font=Font(color=cfg.COLOR_VERDE[1]), fill=PatternFill(start_color=cfg.COLOR_VERDE[0], end_color=cfg.COLOR_VERDE[0], fill_type="solid"))
    s_dos_y = DifferentialStyle(font=Font(color=cfg.COLOR_AMARILLO[1]), fill=PatternFill(start_color=cfg.COLOR_AMARILLO[0], end_color=cfg.COLOR_AMARILLO[0], fill_type="solid"))
    
    st_c = get_column_letter(ci_f + 1)
    for i in lista_reglas:
        r = i["row_actual"]; t = i["row_target"]
        rn = f"{st_c}{r}:{en}{r}"; sc = f"{st_c}{r}"; tc = f"{st_c}{t}"
        ws.conditional_formatting.add(rn, Rule(type="cellIs", operator="greaterThan", formula=[80], dxf=s_dos_p, stopIfTrue=True))
        ws.conditional_formatting.add(rn, Rule(type="cellIs", operator="lessThan", formula=[15], dxf=s_dos_r, stopIfTrue=True))
        ws.conditional_formatting.add(rn, Rule(type="expression", formula=[f"={sc}>{tc}"], dxf=s_dos_g, stopIfTrue=True))
        ws.conditional_formatting.add(rn, Rule(type="cellIs", operator="greaterThanOrEqual", formula=[15], dxf=s_dos_y, stopIfTrue=True))