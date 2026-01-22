import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Listbox, MULTIPLE, Button, Scrollbar, RIGHT, Y, END, LEFT, BOTH, Label, Frame, Radiobutton, IntVar
import os
import pandas as pd
from openpyxl.utils import get_column_letter

# IMPORTACIONES
import app_config as cfg
import utils
import styles
import reports

# --- FUNCIONES GUI ---
def pedir_semanas_bloqueadas(parent, headers, col_inicio, col_fin):
    indices_bloqueados = set()
    ventana = Toplevel(parent); ventana.title("Cierres"); ventana.geometry("400x500")
    Label(ventana, text="Semanas de CIERRE (Ctrl+Click):", padx=10, pady=10).pack()
    frame = Frame(ventana); frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
    sb = Scrollbar(frame); sb.pack(side=RIGHT, fill=Y)
    lb = Listbox(frame, selectmode=MULTIPLE, yscrollcommand=sb.set); lb.pack(side=LEFT, fill=BOTH, expand=True)
    sb.config(command=lb.yview)
    mapa = []
    for c in range(col_inicio + 1, col_fin + 1):
        val = str(headers[c]).split(" ")[0]
        lb.insert(END, val); mapa.append(c)
    def ok():
        for i in lb.curselection(): indices_bloqueados.add(mapa[i])
        ventana.destroy()
    Button(ventana, text="Confirmar", command=ok, bg="#dddddd").pack(fill='x', padx=10, pady=10)
    ventana.wait_window()
    return indices_bloqueados

def menu_seleccion_pdf(root, hubo_simulacion):
    dialogo = Toplevel(root); dialogo.title("Reportes PDF"); dialogo.geometry("350x250")
    Label(dialogo, text="¿Qué reporte PDF deseas generar?", font=("Arial", 11, "bold"), pady=15).pack()
    opcion = IntVar(); opcion.set(1)
    Radiobutton(dialogo, text="1. Reporte de Estado Actual (Formulado)", variable=opcion, value=1, font=("Arial", 10)).pack(anchor="w", padx=20, pady=5)
    rb2 = Radiobutton(dialogo, text="2. Reporte Escenario Ideal", variable=opcion, value=2, font=("Arial", 10))
    rb2.pack(anchor="w", padx=20, pady=5)
    rb3 = Radiobutton(dialogo, text="3. Ambos (Independientes)", variable=opcion, value=3, font=("Arial", 10))
    rb3.pack(anchor="w", padx=20, pady=5)
    if not hubo_simulacion:
        rb2.config(state="disabled"); rb3.config(state="disabled")
        Label(dialogo, text="(Requiere simulación)", fg="red", font=("Arial", 8)).pack(anchor="w", padx=40)
    sel = {"val": 0}
    def ok(): sel["val"] = opcion.get(); dialogo.destroy()
    def no(): sel["val"] = 0; dialogo.destroy()
    fr = Frame(dialogo); fr.pack(fill='x', pady=20)
    Button(fr, text="Generar", command=ok, bg="#4CAF50", fg="white", width=10).pack(side=LEFT, padx=20)
    Button(fr, text="Cancelar", command=no, width=10).pack(side=RIGHT, padx=20)
    dialogo.wait_window()
    return sel["val"]

# --- ESTRUCTURAS DE DATOS ---
def init_report_data():
    return {
        'status_data': [],      
        'heatmap': {},          
        'resumen_cambios': {},  
        'count': 0
    }

# --- ORQUESTADOR PRINCIPAL ---
if __name__ == "__main__":
    root = tk.Tk(); root.withdraw()
    ruta_archivo = filedialog.askopenfilename(title="Selecciona el archivo Excel ORIGINAL", filetypes=[("Archivos Excel", "*.xlsx *.xls")])

    if ruta_archivo and os.path.exists(ruta_archivo):
        nombre_archivo = os.path.basename(ruta_archivo)
        print(f"--- PROCESANDO: {nombre_archivo} ---")
        
        try:
            df_base = pd.read_excel(ruta_archivo, header=None)
            ejecutar_simulacion = messagebox.askyesno("Modo", "¿Ejecutar SIMULACIÓN DEL ESCENARIO IDEAL?\n\nSí: Genera Fórmulas + Ideal\nNo: Solo Fórmulas")
            
            # Limpieza inicial
            for col_idx in range(4): df_base.iloc[2:, col_idx] = df_base.iloc[2:, col_idx].ffill()
            col_inicio = -1; col_fin = -1
            for c in range(df_base.shape[1]):
                val = str(df_base.iloc[0, c])
                if "1900-01-01" in val: col_inicio = c
                if "2222-11-11" in val: col_fin = c
            if col_inicio == -1 or col_fin == -1: raise Exception("No se encontraron fechas clave.")
            
            col_map = {"inicio": col_inicio, "fin": col_fin, "inicio_final": col_inicio + 2, "fin_final": col_fin + 2}
            semanas_cerradas = set()
            if ejecutar_simulacion and messagebox.askyesno("Cierre", "¿Hay semanas de CIERRE de fábrica?"):
                semanas_cerradas = pedir_semanas_bloqueadas(root, df_base.iloc[1].tolist(), col_inicio, col_fin)
            
            productos = df_base.iloc[2:, 2].dropna().unique()
            
            # ---------------------------------------------------------
            # 1. FORMULAS (SITUACIÓN ACTUAL)
            # ---------------------------------------------------------
            print("--- Generando Fórmulas ---")
            df_form = df_base.copy()
            listas_form = {'grupos': [], 'comments': [], 'reglas_dos': [], 'filas_dmd': []}
            
            data_form_global = init_report_data()
            data_form_odm = {} 
            data_form_odm_cat = {}

            count_form = 0
            
            for prod in productos:
                filas = df_form.index[df_form.iloc[:, 2] == prod].tolist()
                ini = min(filas) + 1; fin = max(filas) + 1
                if fin > ini: listas_form['grupos'].append((ini, fin))
                
                odm_val = str(df_form.iloc[filas[0], 1]).strip()
                prefix = str(prod)[:2].upper()
                cat_name = cfg.CATEGORIAS_MAP.get(prefix, "General / Otros")
                key_odm_cat = (odm_val, cat_name)

                if odm_val not in data_form_odm: data_form_odm[odm_val] = init_report_data()
                if key_odm_cat not in data_form_odm_cat: data_form_odm_cat[key_odm_cat] = init_report_data()

                mapa = {utils.normalizar(df_form.iloc[f, 4]): f for f in filas}
                k = {key: utils.normalizar(val) for key, val in cfg.NOMBRES_BUSCADOS.items()}
                if not (k['inv'] in mapa and k['dmd'] in mapa and k['wcd'] in mapa): continue 
                idx = {'inv': mapa[k['inv']], 'dmd': mapa[k['dmd']], 'wcd': mapa[k['wcd']], 'dos_act': mapa.get(k['dos_act']), 'tgt': mapa.get(k['tgt_inv']), 'dos_tgt': mapa.get(k['dos_tgt'])}
                listas_form['filas_dmd'].append(idx['dmd'] + 1)
                if idx['dos_act']: listas_form['reglas_dos'].append({"row_actual": idx['dos_act'] + 1, "row_target": idx['dos_tgt'] + 1})
                
                inv = utils.limpiar_numero(df_form.iloc[idx['inv'], col_inicio]); 
                
                st_excel = ""  
                st_pdf = "OK"  
                
                suma_inv_eval = 0; suma_tgt_eval = 0
                suma_dos_risk = 0; suma_tgt_risk = 0; count_risk = 0
                recovery_week = "-" 
                bad_status_seen = False

                for c in range(col_inicio + 1, col_fin + 1):
                    dmd = utils.limpiar_numero(df_form.iloc[idx['dmd'], c]); wcd = utils.limpiar_numero(df_form.iloc[idx['wcd'], c])
                    inv = inv - dmd + wcd
                    ca = get_column_letter(c + (1 if c >= 4 else 0)); cn = get_column_letter(c + 2 if c + 1 >= 4 else c + 1)
                    df_form.iloc[idx['inv'], c] = f"=ROUND({ca}{idx['inv']+1}-{cn}{idx['dmd']+1}+{cn}{idx['wcd']+1}, 1)"
                    
                    if idx['dos_act']:
                        tgt = utils.limpiar_numero(df_form.iloc[idx['tgt'], c]); dos_t = utils.limpiar_numero(df_form.iloc[idx['dos_tgt'], c])
                        dos = (inv / tgt * dos_t) if tgt else 0
                        df_form.iloc[idx['dos_act'], c] = f"=IFERROR(ROUND(({cn}{idx['inv']+1}/{cn}{idx['tgt']+1})*{cn}{idx['dos_tgt']+1}, 1), 0)"
                        
                        if c < col_inicio + 1 + cfg.SEMANAS_A_EVALUAR:
                            if dmd != 0 and tgt != 0:
                                suma_inv_eval += inv; suma_tgt_eval += tgt
                                
                                # --- EVALUACIÓN COMPLETA PARA PDF ---
                                tipo_st_sem = "OK"
                                if dos <= 0: tipo_st_sem = "OOS"
                                elif dos < dos_t: tipo_st_sem = "USTN"
                                elif dos >= (dos_t + 30): tipo_st_sem = "OSTN High" 
                                elif dos >= (dos_t + 15): tipo_st_sem = "OSTN Med"
                                
                                priority = {"OOS": 1, "USTN": 2, "OSTN High": 3, "OSTN Med": 4, "OK": 5}
                                if priority.get(tipo_st_sem, 5) < priority.get(st_pdf, 5):
                                    st_pdf = tipo_st_sem
                                
                                # GUARDAR EN HEATMAP
                                if tipo_st_sem != "OK":
                                    data_form_global['heatmap'].setdefault(c, []).append((prod, tipo_st_sem))
                                    data_form_odm[odm_val]['heatmap'].setdefault(c, []).append((prod, tipo_st_sem))
                                    data_form_odm_cat[key_odm_cat]['heatmap'].setdefault(c, []).append((prod, tipo_st_sem))

                                # Excel Status & Recovery
                                if tipo_st_sem in ["OOS", "USTN"]:
                                    st_excel = tipo_st_sem 
                                    bad_status_seen = True
                                    if recovery_week == "-": recovery_week = "No Rec."
                                    suma_dos_risk += dos; suma_tgt_risk += dos_t; count_risk += 1
                                
                                elif bad_status_seen and recovery_week == "No Rec." and dos >= dos_t:
                                    header_val = str(df_base.iloc[1, c])
                                    if header_val == 'nan' or header_val == '': header_val = f"Sem {c - col_inicio}"
                                    recovery_week = header_val

                gap_total = suma_tgt_eval - suma_inv_eval 
                if count_risk > 0: avg_dos_a = suma_dos_risk / count_risk; avg_dos_t = suma_tgt_risk / count_risk
                else: avg_dos_a = 0; avg_dos_t = 0
                
                info_prod_pdf = {
                    "row": ini, "status": st_pdf, "gap": gap_total, 
                    "avg_dos_act": avg_dos_a, "avg_dos_tgt": avg_dos_t, "recovery": recovery_week 
                }
                
                listas_form['comments'].append({"row": ini, "status": st_excel}) 
                data_form_global['status_data'].append(info_prod_pdf)
                data_form_global['count'] += 1
                data_form_odm[odm_val]['status_data'].append(info_prod_pdf)
                data_form_odm[odm_val]['count'] += 1
                data_form_odm_cat[key_odm_cat]['status_data'].append(info_prod_pdf)
                data_form_odm_cat[key_odm_cat]['count'] += 1
                
                count_form += 1

            if count_form > 0:
                ruta = filedialog.asksaveasfilename(title="Guardar FÓRMULAS", defaultextension=".xlsx", initialfile=f"Formulado_{nombre_archivo}")
                if ruta:
                    with pd.ExcelWriter(ruta, engine='openpyxl') as w: 
                        df_form.to_excel(w, index=False, header=False, sheet_name='Sheet1')
                        styles.aplicar_maquillaje(w.sheets['Sheet1'], df_form, listas_form, col_map)

            # ---------------------------------------------------------
            # 2. IDEAL (SIMULACIÓN + REPORTE SEPARADO)
            # ---------------------------------------------------------
            listas_ideal = None
            data_ideal_global = init_report_data(); data_ideal_odm = {}; data_ideal_odm_cat = {}
            count_ideal = 0; df_ideal = None
            
            if ejecutar_simulacion and count_form > 0:
                print("--- Generando Ideal ---")
                df_ideal = df_base.copy()
                listas_ideal = {'grupos': [], 'comments': [], 'reglas_dos': [], 'filas_dmd': []}
                for prod in productos:
                    filas = df_ideal.index[df_ideal.iloc[:, 2] == prod].tolist()
                    ini = min(filas) + 1; fin = max(filas) + 1
                    if fin > ini: listas_ideal['grupos'].append((ini, fin))
                    
                    odm_val = str(df_ideal.iloc[filas[0], 1]).strip()
                    prefix = str(prod)[:2].upper()
                    cat_name = cfg.CATEGORIAS_MAP.get(prefix, "General / Otros")
                    key_odm_cat = (odm_val, cat_name)

                    if odm_val not in data_ideal_odm: data_ideal_odm[odm_val] = init_report_data()
                    if key_odm_cat not in data_ideal_odm_cat: data_ideal_odm_cat[key_odm_cat] = init_report_data()

                    mapa = {utils.normalizar(df_ideal.iloc[f, 4]): f for f in filas}
                    k = {key: utils.normalizar(val) for key, val in cfg.NOMBRES_BUSCADOS.items()}
                    if not (k['inv'] in mapa and k['dmd'] in mapa and k['wcd'] in mapa): continue
                    idx = {'inv': mapa[k['inv']], 'dmd': mapa[k['dmd']], 'wcd': mapa[k['wcd']], 'wsd': mapa.get(k['wsd']), 'dos_act': mapa.get(k['dos_act']), 'tgt': mapa.get(k['tgt_inv']), 'dos_tgt': mapa.get(k['dos_tgt'])}
                    listas_ideal['filas_dmd'].append(idx['dmd'] + 1); listas_ideal['reglas_dos'].append({"row_actual": idx['dos_act'] + 1, "row_target": idx['dos_tgt'] + 1})
                    
                    LT = cfg.LEAD_TIMES.get(str(df_ideal.iloc[ini-1, 0]), 8)
                    
                    # === FASE 1: OPTIMIZACIÓN (MOVIMIENTOS) ===
                    for _ in range(500):
                        bad = False; inv = utils.limpiar_numero(df_ideal.iloc[idx['inv'], col_inicio])
                        for c in range(col_inicio + 1, col_fin + 1):
                            dmd = utils.limpiar_numero(df_ideal.iloc[idx['dmd'], c])
                            wcd = utils.limpiar_numero(df_ideal.iloc[idx['wcd'], c])
                            inv = inv - dmd + wcd
                            tgt = utils.limpiar_numero(df_ideal.iloc[idx['tgt'], c])
                            dos_t = utils.limpiar_numero(df_ideal.iloc[idx['dos_tgt'], c])
                            dos = (inv / tgt * dos_t) if tgt else 0
                            
                            # Actualizar valores en DataFrame para la siguiente iteración
                            df_ideal.iloc[idx['inv'], c] = round(inv, 1)
                            df_ideal.iloc[idx['dos_act'], c] = round(dos, 1)
                            
                            if c >= col_inicio + LT + 2 and c < col_inicio + LT + 2 + cfg.SEMANAS_A_EVALUAR:
                                if dmd != 0 and tgt != 0 and dos < dos_t and dos_t > 0:
                                    dias_objetivo = dos_t + 5
                                    need = (tgt * dias_objetivo) / dos_t; delta = round(need - inv, 0)
                                    if delta > 0:
                                        wsd_c = c - LT; bad = True 
                                        if wsd_c in semanas_cerradas: break
                                        movs = []
                                        encontrado = False
                                        limit_semanas = wsd_c + cfg.MAX_SEMANAS_RECORRER + 1
                                        if limit_semanas > col_fin: limit_semanas = col_fin
                                        
                                        def get_header(col_idx):
                                            val = str(df_ideal.iloc[1, col_idx])
                                            if val == 'nan' or val.strip() == '': val = str(df_ideal.iloc[0, col_idx]).split(" ")[0]
                                            return val

                                        for cs in range(wsd_c + 1, limit_semanas):
                                            ex = utils.limpiar_numero(df_ideal.iloc[idx['wsd'], cs])
                                            if ex > 0:
                                                df_ideal.iloc[idx['wsd'], wsd_c] += ex; df_ideal.iloc[idx['wsd'], cs] = 0
                                                df_ideal.iloc[idx['wcd'], c] += ex; df_ideal.iloc[idx['wcd'], cs + LT] -= ex
                                                movs.append({"tipo": "recorrido", "cantidad": ex, "col_origen": cs, "col_destino": wsd_c, "str_origen": get_header(cs), "str_destino": get_header(wsd_c)})
                                                encontrado = True; break
                                        
                                        if not encontrado:
                                            rem = delta; df_ideal.iloc[idx['wcd'], c] += rem; df_ideal.iloc[idx['wsd'], wsd_c] += rem
                                            movs.append({"tipo": "aumento", "cantidad": rem, "col_destino": wsd_c, "str_destino": get_header(wsd_c)})

                                        if movs: 
                                            data_ideal_global['resumen_cambios'].setdefault(prod, []).extend(movs)
                                            data_ideal_odm[odm_val]['resumen_cambios'].setdefault(prod, []).extend(movs)
                                            data_ideal_odm_cat[key_odm_cat]['resumen_cambios'].setdefault(prod, []).extend(movs)
                                    break
                        if not bad: break
                    
                    # === FASE 2: REPORTE (CLASIFICACIÓN FINAL) ===
                    # Recorremos una última vez el producto "optimizado" para clasificarlo
                    inv = utils.limpiar_numero(df_ideal.iloc[idx['inv'], col_inicio])
                    
                    st_excel = ""; st_pdf = "OK"
                    suma_inv_eval = 0; suma_tgt_eval = 0
                    suma_dos_risk = 0; suma_tgt_risk = 0; count_risk = 0
                    recovery_week = "-"; bad_status_seen = False

                    for c in range(col_inicio + 1, col_fin + 1):
                        dmd = utils.limpiar_numero(df_ideal.iloc[idx['dmd'], c])
                        wcd = utils.limpiar_numero(df_ideal.iloc[idx['wcd'], c])
                        inv = inv - dmd + wcd
                        tgt = utils.limpiar_numero(df_ideal.iloc[idx['tgt'], c])
                        dos_t = utils.limpiar_numero(df_ideal.iloc[idx['dos_tgt'], c])
                        dos = (inv / tgt * dos_t) if tgt else 0
                        
                        if c < col_inicio + 1 + cfg.SEMANAS_A_EVALUAR:
                            if dmd != 0 and tgt != 0:
                                suma_inv_eval += inv; suma_tgt_eval += tgt
                                
                                # PDF Status Check
                                tipo_st_sem = "OK"
                                if dos <= 0: tipo_st_sem = "OOS"
                                elif dos < dos_t: tipo_st_sem = "USTN"
                                elif dos >= (dos_t + 30): tipo_st_sem = "OSTN High"
                                elif dos >= (dos_t + 15): tipo_st_sem = "OSTN Med"
                                
                                priority = {"OOS": 1, "USTN": 2, "OSTN High": 3, "OSTN Med": 4, "OK": 5}
                                if priority.get(tipo_st_sem, 5) < priority.get(st_pdf, 5): st_pdf = tipo_st_sem

                                # GUARDAR EN HEATMAP (Cualquier desviación)
                                if tipo_st_sem != "OK":
                                    data_ideal_global['heatmap'].setdefault(c, []).append((prod, tipo_st_sem))
                                    data_ideal_odm[odm_val]['heatmap'].setdefault(c, []).append((prod, tipo_st_sem))
                                    data_ideal_odm_cat[key_odm_cat]['heatmap'].setdefault(c, []).append((prod, tipo_st_sem))

                                # Excel Status & Recovery Logic
                                if dos < dos_t:
                                    bad_status_seen = True; st_excel = "OOS" if dos <= 0 else "USTN"
                                    if recovery_week == "-": recovery_week = "No Rec."
                                    suma_dos_risk += dos; suma_tgt_risk += dos_t; count_risk += 1
                                        
                                elif bad_status_seen and recovery_week == "No Rec." and dos >= dos_t:
                                    header_val = str(df_base.iloc[1, c])
                                    if header_val == 'nan' or header_val == '': header_val = f"Sem {c - col_inicio}"
                                    recovery_week = header_val

                    gap_total = suma_tgt_eval - suma_inv_eval
                    if count_risk > 0: avg_dos_a = suma_dos_risk / count_risk; avg_dos_t = suma_tgt_risk / count_risk
                    else: avg_dos_a = 0; avg_dos_t = 0
                    
                    listas_ideal['comments'].append({"row": ini, "status": st_excel}) # Excel solo OOS/USTN
                    
                    info_prod_pdf = {
                        "row": ini, "status": st_pdf, "gap": gap_total, 
                        "avg_dos_act": avg_dos_a, "avg_dos_tgt": avg_dos_t, "recovery": recovery_week
                    }
                    data_ideal_global['status_data'].append(info_prod_pdf)
                    data_ideal_global['count'] += 1
                    data_ideal_odm[odm_val]['status_data'].append(info_prod_pdf)
                    data_ideal_odm[odm_val]['count'] += 1
                    data_ideal_odm_cat[key_odm_cat]['status_data'].append(info_prod_pdf)
                    data_ideal_odm_cat[key_odm_cat]['count'] += 1
                    count_ideal += 1

                if count_ideal > 0:
                    ruta = filedialog.asksaveasfilename(title="Guardar IDEAL", defaultextension=".xlsx", initialfile=f"Escenario_Ideal_{nombre_archivo}")
                    if ruta:
                        with pd.ExcelWriter(ruta, engine='openpyxl') as w:
                            df_ideal.to_excel(w, index=False, header=False, sheet_name='Sheet1')
                            styles.aplicar_maquillaje(w.sheets['Sheet1'], df_ideal, listas_ideal, col_map)

            if reports.LIBRERIA_PDF_DISPONIBLE and count_form > 0:
                eleccion = menu_seleccion_pdf(root, ejecutar_simulacion)
                lt_func = lambda x: cfg.LEAD_TIMES.get(x, 8)
                paquete_actual = {'global': data_form_global, 'odm': data_form_odm, 'odm_cat': data_form_odm_cat}
                paquete_ideal = {'global': data_ideal_global, 'odm': data_ideal_odm, 'odm_cat': data_ideal_odm_cat}
                
                if eleccion == 1 or eleccion == 3:
                    ruta = filedialog.asksaveasfilename(title="Guardar PDF (Actual)", defaultextension=".pdf", initialfile=f"Reporte_Inicial_{nombre_archivo}.pdf")
                    if ruta: reports.generar_reporte_pdf(ruta, df_base, lt_func, "FORMULAS", data_pack_main=paquete_actual, col_inicio=col_inicio)
                
                if eleccion == 2 or eleccion == 3:
                    ruta = filedialog.asksaveasfilename(title="Guardar PDF (Ideal)", defaultextension=".pdf", initialfile=f"Reporte_Ideal_{nombre_archivo}.pdf")
                    if ruta and listas_ideal: reports.generar_reporte_pdf(ruta, df_base, lt_func, "IDEAL", data_pack_main=paquete_ideal, data_pack_antes=paquete_actual, col_inicio=col_inicio)
                messagebox.showinfo("Listo", "Proceso Finalizado.")
            else:
                if not reports.LIBRERIA_PDF_DISPONIBLE: messagebox.showwarning("Aviso", "No se encontró reportlab.")
                else: messagebox.showinfo("Listo", "Proceso Finalizado.")

        except Exception as e: messagebox.showerror("Error", str(e))
    else: root.destroy()